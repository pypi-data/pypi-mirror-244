"""Excel workbook and worksheets utilities."""
import os
import sys
from openpyxl import load_workbook, utils as openpyxl_utils
import zipfile
from termcolor import cprint
import colorama
import asyncio

colorama.init()

ERROR_COLOUR = 'red'
INVALID_EXCEL_FILE = 'is not a valid excel file'
QUITTING = 'Quitting ...'
PATH_ERROR = 'File path error'
WORKBOOK_ERROR = 'Workbook error'
WORKSHEET_ERROR = 'Worksheet error'
WORKBOOK_MISSING = 'Cannot find workbook:'
NAMED_CORRECTLY = 'Is it named correctly?'
MISSING_SHEET = ': missing sheet in the workbook'


class Workbook():
    def __init__(self, workbook_path: str):
        self.path = str(workbook_path)
        self.workbook = self._get_workbook()
        self.worksheets = {sheet._WorkbookChild__title: sheet for sheet in self.workbook.worksheets}

    def _get_workbook(self) -> object:
        """Return the workbook if it exists,raise an error. """
        if not os.path.isfile(self.path):
            cprint(f"{WORKBOOK_MISSING} {self.path}", ERROR_COLOUR)
            cprint(f'{NAMED_CORRECTLY}', ERROR_COLOUR)
            # sys.exit(f'{QUITTING} {PATH_ERROR}')
            raise FileNotFoundError()
        try:
            workbook = load_workbook(filename=self.path, data_only=True)
            return workbook
        except openpyxl_utils.exceptions.InvalidFileException:
            cprint(f"{self.path} {INVALID_EXCEL_FILE}", ERROR_COLOUR)
            sys.exit(f'{QUITTING} {WORKBOOK_ERROR}')
        except zipfile.BadZipFile:
            cprint(f"{self.path} {INVALID_EXCEL_FILE}", ERROR_COLOUR)
            sys.exit(f'{QUITTING} {WORKBOOK_ERROR}')

    def list_worksheets(self) -> None:
        "Print worksheet names."
        print('')
        print('*'*5, 'Sheets', '*'*5)
        for sheet in self.worksheets:
            print(sheet._WorkbookChild__title)
        print('*'*18)

    async def get_worksheet(self, sheet_name: str) -> object:
        """Return the worksheet if it exists, otherwise quit."""
        try:
            worksheet =  self.workbook[sheet_name]
        except KeyError:
            worksheet = None
            msg = f"{sheet_name}{MISSING_SHEET} {self.path}"
            cprint(msg, ERROR_COLOUR)
            raise
            # sys.exit(f'{QUITTING} {WORKSHEET_ERROR}')
        return worksheet

    def rows_from_worksheet(self, sheet_name, ignore_first_row=True):
        """Return all of the rows from a worksheet."""
        worksheet = self.worksheets[sheet_name]
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(row)
        if ignore_first_row:
            rows.pop(0)
        return rows

    def save(self, path=None) -> None:
        """Save the workbook to the path."""
        if not path:
            path = self.path
        self.workbook.save(path)


class Worksheet():
    """Class to mirror openpyxl worksheet. Only use for type documentation"""
    def __init__(self):
        ...
