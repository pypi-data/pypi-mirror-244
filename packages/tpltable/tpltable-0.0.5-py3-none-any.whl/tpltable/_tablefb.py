import math
import copy
import time
import sys
import os
from typing import Union

# from filefinder import *

# Handle Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
# openpyxl.worksheet.worksheet.Worksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def load_workbook(fpath: str):
    """
    加载工作簿
    :param fpath:
    :return:
    """
    wb = openpyxl.load_workbook(fpath)
    fnametype = os.path.basename(fpath)
    wb.name = fnametype
    return wb


class TableFB:  # Table Formatter Backend
    @staticmethod
    def _get_bookinfo(book: Workbook, justify=True) -> str:
        _ = getattr(book, 'name', f"Unknown-WorkBook")
        _ = f"[{_}]"
        if justify:
            _ = _.ljust(20)
        return _

    @staticmethod
    def _get_sheetinfo(sheet: Worksheet, justify=True) -> str:
        parent = getattr(sheet, "parent", None)
        _p = getattr(parent, 'name', f"Unknown-WorkBook")
        _t = getattr(sheet, 'title', f"Unknown-Sheet")
        _ = f"[{_p}.{_t}]"
        if justify:
            _ = _.ljust(25)
        return _

    @staticmethod
    def _calculate_cell_count(sheet):
        cell_count = 0
        merged_cells_coordinates = set()
        for merged in sheet.merged_cells.ranges:
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    coordinate = get_column_letter(col) + str(row)
                    merged_cells_coordinates.add(coordinate)
            cell_count += int(math.sqrt((merged.max_row - merged.min_row + 1) * (merged.max_col - merged.min_col + 1)))
        for row in sheet.iter_rows():
            for cell in row:
                if cell.coordinate not in merged_cells_coordinates:
                    cell_count += 1
        return cell_count

    @staticmethod
    def _crop_offset(ws: Worksheet, offset: list) -> Worksheet:
        """
        移除目标sheet的前offset列和行
        """
        if offset[1] > 0:
            ws.delete_rows(1, offset[1])
        if offset[0] > 0:
            ws.delete_cols(1, offset[0])
        return ws

    @staticmethod
    def _try_show_progress(progress:float, _next_time:float):
        # print(time.time(), _next_time)
        if time.time() >= _next_time:
            progress *= 100

            # 像这样: [=====>       ] 30%
            # 一共20个等号, 100%的时候是20个等号
            equal_str = '=' * int(progress // 5)
            blank_str = ' ' * (20 - int(progress // 5))
            print(f"[{equal_str + '>' + blank_str}] {progress:.1f}%", end='\r')
            # sys.stdout.flush()
            return True
        return False

    @staticmethod
    def _scan_table(sheet: Worksheet):
        """
        扫描单个sheet, 记录以$开头的单元格(包括合并格)的位置, 以及对应的值
        :param sheet: openpyxl.worksheet.worksheet.Worksheet
        :return:
        """
        tdata = {}
        for merge in sheet.merged_cells.ranges:
            if merge.start_cell.value and str(merge.start_cell.value).startswith('$'):
                tdata[str(merge)] = merge.start_cell.value

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('$') and cell.coordinate not in tdata:
                    tdata[cell.coordinate] = cell.value

        # 检查values有无重复, 若重复, 报错并指出两个重复值的key
        vk_dict = {}
        for key, val in tdata.items():
            if val not in vk_dict:
                vk_dict[val] = [key]
            elif len(vk_dict[val]) == 1 and (':' in key) == (':' in vk_dict[val][0]):
                raise ValueError(f"values have duplicates: {val} in {key} and {vk_dict[val][0]}")
            elif len(vk_dict[val]) >= 2:
                raise ValueError(f"values have duplicates: {val} in {key} and {vk_dict[val][0]}, {vk_dict[val][1]}")

        return tdata

    @staticmethod
    def _match_table(sheet_template: Worksheet, target_sheet: Worksheet, patience_count: int, debug=False):
        """
        比较两个sheet的内容是否一致, 如果不一致, 返回差异
        :param sheet_template:
        :param target_sheet:
        :param patience_count: 容忍的差异个数
        :param debug: debug时返回str而不是bool
        :return:
        """
        differences = []

        # 存储所有合并单元格的坐标
        merged_cells_coordinates = set()
        for merged in sheet_template.merged_cells.ranges:
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    coordinate = get_column_letter(col) + str(row)
                    merged_cells_coordinates.add(coordinate)

        # 比对合并单元格
        for merged in sheet_template.merged_cells.ranges:
            if str(merged) not in target_sheet.merged_cells:
                differences.append(f"$合并单元格 {merged} 在目标表格中不存在")
            # else:
            #     # 检查合并单元格的内容是否一致
            #     top_left_cell = sheet_template[merged.min_row][merged.min_col]
            #     target_top_left_cell = target_sheet[top_left_cell.coordinate]
            #     if top_left_cell.value != target_top_left_cell.value:
            #         differences.append(f"合并单元格 {merged} 的内容不一致. 模板: {top_left_cell.value}, 目标: {target_top_left_cell.value}")

        # 比对单元格内容
        for row in sheet_template.iter_rows():
            for cell in row:
                if not str(cell.value).startswith("$") and cell.coordinate not in merged_cells_coordinates:
                    target_cell = target_sheet[cell.coordinate]
                    if cell.value != target_cell.value:
                        differences.append(f"单元格 {cell.coordinate} 的内容不一致. 模板: {cell.value}, 目标: {target_cell.value}")

        if len(differences) > patience_count:
            return (f"差异过大: {len(differences)} > {patience_count}:\n" + '\n'.join(differences)) if debug else False
        return True

    @staticmethod
    def _is_empty_row(sheet, irow):
        """
        判断一行是否为空
        :param sheet:
        :param irow:
        :return:
        """
        for cell in sheet[irow]:
            if cell.value:
                return False
        return True

    @staticmethod
    def _is_empty_col(sheet, icol):
        """
        判断一列是否为空
        :param sheet:
        :param icol:
        :return:
        """
        for cell in sheet[icol]:
            if cell.value:
                return False
        return True

    @staticmethod
    def _calculate_template_shape(sheet, fDATA):
        """
        计算fDATA的长和宽
        """
        min_row, max_row, min_col, max_col = float('inf'), 0, float('inf'), 0
        for key in fDATA.keys():
            _min_col, _min_row, _max_col, _max_row = range_boundaries(key)
            min_row, max_row = min(min_row, _min_row), max(max_row, _max_row)
            min_col, max_col = min(min_col, _min_col), max(max_col, _max_col)

        # 逼近min_row和min_col
        for i in range(min_row - 1, 0, -1):
            if not TableFB._is_empty_row(sheet, i):
                min_row = i
                break
        for i in range(min_col - 1, 0, -1):
            if not TableFB._is_empty_col(sheet, i):
                min_col = i
                break

        height = max_row - min_row + 1
        width = max_col - min_col + 1
        return width, height

    @staticmethod
    def _calculate_template_range4(sheet, fDATA):
        """
        计算fDATA的range4
        """
        min_row, max_row, min_col, max_col = float('inf'), 0, float('inf'), 0
        for key in fDATA.keys():
            _min_col, _min_row, _max_col, _max_row = range_boundaries(key)
            min_row, max_row = min(min_row, _min_row), max(max_row, _max_row)
            min_col, max_col = min(min_col, _min_col), max(max_col, _max_col)

        # 逼近min_row和min_col
        for i in range(min_row - 1, 0, -1):
            if not TableFB._is_empty_row(sheet, i):
                min_row = i
                break
        for i in range(min_col - 1, 0, -1):
            if not TableFB._is_empty_col(sheet, i):
                min_col = i
                break

        return min_row, max_row, min_col, max_col

    @staticmethod
    def _move_template(fDATAr, n, m):
        """
        "移动"模板fDATAr
        x方向移动n个长度,y方向移动m个宽度
        """
        new_fDATAr = {}
        for key, value in fDATAr.items():
            _min_row, _max_row, _min_col, _max_col = key
            new_key = (_min_row + m, _max_row + m, _min_col + n, _max_col + n)
            new_fDATAr[new_key] = value
        return new_fDATAr

    @staticmethod
    def _convert_keys_to_range(fDATA):
        """
        将fDATA中的key转换为range4: (min_row, max_row, min_col, max_col)
        返回一个新的fDATAr
        """
        fDATAr = {}
        for key, value in fDATA.items():
            _min_col, _min_row, _max_col, _max_row = range_boundaries(key)
            new_key = (_min_row, _max_row, _min_col, _max_col)
            fDATAr[new_key] = value
        return fDATAr

    @staticmethod
    def _copy_style(src_cell, dest_cell):
        """
        复制源单元格的样式到目标单元格。
        :param src_cell: 源单元格
        :param dest_cell: 目标单元格
        """
        if src_cell.has_style:
            dest_cell.font = copy.copy(src_cell.font)
            dest_cell.border = copy.copy(src_cell.border)
            dest_cell.fill = copy.copy(src_cell.fill)
            dest_cell.number_format = copy.copy(src_cell.number_format)
            dest_cell.protection = copy.copy(src_cell.protection)
            dest_cell.alignment = copy.copy(src_cell.alignment)

    @staticmethod
    def _iterate_over_target(target: Worksheet, fDATA, force_shape):
        """
        迭代器，输入target:WorkSheet, fDATA, force_shape:tuple (col, row)
        每次yield一个新的WorkSheet, 这个新的ws只有模板的长宽, 按照先行后列的顺序在target上"移动"模板
        """
        fDATAr = TableFB._convert_keys_to_range(fDATA)
        template_width, template_height = TableFB._calculate_template_shape(target, fDATA)
        rows, cols = target.max_row, target.max_column
        template_ix = 0
        template_iy = 0

        if force_shape is None:
            # Auto Scan Mode
            range_row = range(1, rows + 1, template_height)
            range_col = range(1, cols + 1, template_width)
        else:
            # Force Shape Mode
            range_row = range(1, template_height * force_shape[0] - 1, template_height)
            range_col = range(1, template_width * force_shape[1] - 1, template_width)
            # 补全目标:
            # 如果目标的长宽不足, 则在尾部新增空白行/列
            if rows < template_height * force_shape[0]:
                target.insert_rows(rows + 1, template_height * force_shape[0] - rows)
            if cols < template_width * force_shape[1]:
                target.insert_cols(cols + 1, template_width * force_shape[1] - cols)
            rows, cols = max(rows, template_height * force_shape[0]), max(cols, template_width * force_shape[1])

        for i in range_row:  # 第一个维度是行(对应y轴)
            for j in range_col:  # 第二个维度是列(对应x轴)
                if i + template_height - 1 > rows or j + template_width - 1 > cols:
                    continue
                wb = Workbook()
                ws = wb.active
                for row in range(i, i + template_height):
                    for col in range(j, j + template_width):
                        # value font border fill
                        cell = target.cell(row=row, column=col)
                        new_cell = ws.cell(row=row - i + 1, column=col - j + 1)
                        new_cell.value = cell.value
                        TableFB._copy_style(cell, new_cell)

                # 获取目标区域内的所有合并单元格
                for merged in target.merged_cells.ranges:
                    if merged.min_row >= i and merged.max_row <= i + template_height - 1 and merged.min_col >= j and merged.max_col <= j + template_width - 1:
                        ws.merge_cells(start_row=merged.min_row - i + 1, start_column=merged.min_col - j + 1,
                                       end_row=merged.max_row - i + 1, end_column=merged.max_col - j + 1)

                if template_ix >= cols // template_width:
                    template_ix = 0
                    template_iy += 1

                yield (template_ix, template_iy), (i, i + template_height - 1, j, j + template_width - 1), ws
                template_ix += 1

    @staticmethod
    def _copy_from_to(source: Worksheet, target: Worksheet, source_range4: tuple, target_range4: tuple):
        """
        将lib中的数据写入到target中
        :param source: 源sheet
        :param target: 目标sheet
        :param source_range4: (min_row, max_row, min_col, max_col)
        :param target_range4: (min_row, max_row, min_col, max_col)
        :return:
        """
        assert len(source_range4) == 4, f"source_range4 must be a tuple of 4 elements. But got {len(source_range4)} elements."
        assert len(target_range4) == 4, f"target_range4 must be a tuple of 4 elements. But got {len(target_range4)} elements."
        assert source_range4[1] - source_range4[0] == target_range4[1] - target_range4[0], \
            f"source_range4 and target_range4 must have the same height. But got {source_range4[1] - source_range4[0]} and {target_range4[1] - target_range4[0]}"

        for i in range(source_range4[0], source_range4[1] + 1):
            for j in range(source_range4[2], source_range4[3] + 1):
                src_cell = source.cell(row=i, column=j)
                dest_cell = target.cell(row=i + target_range4[0] - source_range4[0], column=j + target_range4[2] - source_range4[2])
                dest_cell.value = src_cell.value
                TableFB._copy_style(src_cell, dest_cell)

        # Copy merged cells
        for merged_cell_range in source.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if (min_row >= source_range4[0] and max_row <= source_range4[1] and
                    min_col >= source_range4[2] and max_col <= source_range4[3]):
                target.merge_cells(
                    start_row=min_row + target_range4[0] - source_range4[0],
                    start_column=min_col + target_range4[2] - source_range4[2],
                    end_row=max_row + target_range4[0] - source_range4[0],
                    end_column=max_col + target_range4[2] - source_range4[2]
                )

    @staticmethod
    def _paste_lib_to(target: Worksheet, range4: tuple, lib: dict):
        """
        将lib中的数据写入到target中
        :param target: 目标sheet
        :param range4: (min_row, max_row, min_col, max_col)
        :param lib: dict, key为以$开头的字符串, value为值. 当对应单元格的数据为$xxx时, 将其替换为value
        :return:
        """
        assert len(range4) == 4, f"range4 must be a tuple of 4 elements. But got {len(range4)} elements."

        for i in range(range4[0], range4[1] + 1):
            for j in range(range4[2], range4[3] + 1):
                cell = target.cell(row=i, column=j)
                if isinstance(cell.value, str) and cell.value.startswith('$'):
                    if cell.value in lib:
                        cell.value = lib[cell.value]
                    else:
                        print(f"Warning: {cell.value} not found in lib.")


class BookTableFB:

    @staticmethod
    def _check_fpath(excel_fpath: str, error=True):
        if not os.path.exists(excel_fpath):
            if error:
                raise FileNotFoundError(f"File not found: {excel_fpath}")
            else:
                return False
        if not os.path.isfile(excel_fpath):
            if error:
                raise TypeError(f"{excel_fpath} is not a file.")
            else:
                return False
        if excel_fpath[-4:].lower() not in ('.xls', 'xlsx'):
            if error:
                raise TypeError(f"{excel_fpath} is not a excel file.")
            else:
                return False
        return True

    @staticmethod
    def _is_mixed_target(target: list) -> bool:
        """
        检查目标对象是否是混合对象列表
        :param target: list of Workbook|Worksheet
        :return: bool
        """
        types = set([type(t) for t in target])
        return Workbook in types and Worksheet in types

    @staticmethod
    def _get_list_target_type(target: list) -> type:
        """
        返回目标列表中元素的类型
        :param target: list of Workbook|Worksheet
        :return: type
        """
        if not BookTableFB._is_mixed_target(target):
            return type(target[0])
        else:
            raise ValueError(f"目标对象类型不正确，应为Workbook, Worksheet, 或 list of Worksheet|Workbook. But got list of mixed type.")

    @staticmethod
    def _expand_target2sheets(target: Union[Workbook, Worksheet, list]) -> list:
        """
        将目标对象转换为list of Worksheet
        :param target: 可以为Workbook, Worksheet, list of Worksheet|Workbook
        :return: list
        """
        if isinstance(target, Workbook):
            return target.worksheets
        elif isinstance(target, Worksheet):
            return [target]
        elif isinstance(target, (list, tuple)):
            sheets = []
            for t in target:
                if isinstance(t, Workbook):
                    sheets.extend(t.worksheets)
                elif isinstance(t, Worksheet):
                    sheets.append(t)
            return sheets
        else:
            raise ValueError("目标对象类型不正确，应为Workbook, Worksheet, 或 list of Worksheet|Workbook")

    @staticmethod
    def _get_wb_name(wb: Workbook) -> str:
        """
        获取Workbook的名称
        :param wb:
        :return:
        """
        name = getattr(wb, 'name', None)
        if name is None:
            name = wb.name = f"WorkBook<{id(wb)}>"
        return name
