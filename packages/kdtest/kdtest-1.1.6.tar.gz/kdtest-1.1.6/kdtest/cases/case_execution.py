'''# Third_party'''
from openpyxl import load_workbook
'''# built_in'''
import os, importlib, traceback, configparser
'''# custom'''
from common import GSDSTORE, MODULEDATA, PROJECTROOT, CASESDATA, CASES, ST
from cases.read import excel_Read
from Interface.common_Interface_Class import CommonInterfaceClass
from utils.log.trusteeship import trusteeship


class common(object):
    @classmethod
    def restore(cls): trusteeship.test_end(); GSDSTORE['driver'].quit(); quit()


class collector(object):
    @classmethod
    def run(cls):
        trusteeship.test_start(); trusteeship.frame_init()  # 运行日志初始化
        '''Use case data capture'''
        for key in ['total','prepare','implement','passNum','abnormalNum','failNum','suite_setup','suite_teardown','plugin']: CASESDATA[key] = 0
        CASESDATA['toSheetNames'] = []
        try:
            for cases_Json in GSDSTORE['START']['testCaseFile']:
                WB_Object = load_workbook(cases_Json['caseFilePath'], read_only=True)
                toSheetNames = []
                if 'caseItem' in cases_Json.keys() or 'caseStart' in cases_Json.keys() or 'caseEnd' in cases_Json.keys():
                    if 'caseItem' in cases_Json.keys() and cases_Json['caseItem']:  # 指定用例
                        toSheetNames = [item.strip() for item in cases_Json['caseItem'] if WB_Object.sheetnames.index(item.strip()) + 1]
                    else:  # 区间
                        toSheetNames = WB_Object.sheetnames[:]
                        if 'caseStart' in cases_Json.keys() and cases_Json['caseStart'].strip(): toSheetNames = toSheetNames[toSheetNames.index(cases_Json['caseStart'].strip()):]
                        if 'caseEnd' in cases_Json.keys() and cases_Json['caseEnd'].strip(): toSheetNames = toSheetNames[:toSheetNames.index(cases_Json['caseEnd'].strip()) + 1]
                else:
                    toSheetNames = WB_Object.sheetnames[:]
                CASESDATA['toSheetNames'].append(toSheetNames); CASESDATA['total'] += len(WB_Object.sheetnames); CASESDATA['prepare'] += len(toSheetNames); WB_Object.close()
        except:
            trusteeship.exceptinfo("Use cases catch exceptions: ");common.restore()
        '''Initialize the clear capture'''
        try:
            if os.path.exists(f"{GSDSTORE['WORKPATH']['ROOT']}\\{ST}.py"):
                spec = importlib.util.spec_from_file_location('ST', f"{GSDSTORE['WORKPATH']['ROOT']}\\{ST}.py")
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                cls.suite_setup = getattr(module, 'suite_setup')
                cls.suite_teardown = getattr(module, 'suite_teardown')
        except:
            trusteeship.exceptinfo("Initialize to clear caught exceptions: ");common.restore()
        '''Capture plug-in'''
        try:
            cfp = configparser.ConfigParser()
            pluginList = os.listdir(f"{PROJECTROOT}\\plugins")
            for plugin in pluginList:
                moduleini = [py for py in os.listdir(f"{PROJECTROOT}\\plugins\\{plugin}") if os.path.splitext(f"{PROJECTROOT}\\plugins\\{plugin}\\{py}")[1] == ".ini" ][0]
                cfp.read(f"{PROJECTROOT}\\plugins\\{plugin}\\{''.join(moduleini)}", encoding='utf-8')
                name = cfp.get('Information', 'NAME'); detail = cfp.get('Information', 'DETAILE'); state = cfp.get('Information', 'STATE')
                module_Class = getattr(importlib.import_module(f"plugins.{plugin}.{name}"), f'{name}')  # 动态导入
                MODULEDATA[name] = {'information':{}, 'object': object}
                MODULEDATA[name]['information']['name'] = name
                MODULEDATA[name]['information']['detail'] = detail
                MODULEDATA[name]['information']['state'] = state
                MODULEDATA[name]['object'] = module_Class()
        except:
            CASESDATA['plugin'] += 1;trusteeship.exceptinfo("Custom plug-ins catch exceptions: ");common.restore()


class cases(object):
    def __repr__(self) -> str:
        return "测试用例执行控制脚本/逻辑用例脚本"

    def __init__(self) -> None:
        '''object对象'''
        self.__commonInterface_Object = CommonInterfaceClass()

    def __call__(self) -> None:
        '''初始化'''
        self.__initializeClear("suite_setup", trusteeship.case_init())
        
        '''用例'''
        elementFolderPath = None
        for loopIndex, caseInformation in enumerate(GSDSTORE['START']['testCaseFile']):
            '''目录结构追踪'''
            thisFolderPaht = caseInformation['caseFilePath'].split('\\')[:-1]
            subFolderPath = "\\".join(thisFolderPaht[thisFolderPaht.index(CASES) + 1:])

            '''接口'''
            if ('INTERFACE' in GSDSTORE['WORKPATH']) and ('interfaceSwitch' in caseInformation.keys() and caseInformation['interfaceSwitch'] == True):
                interfacePath = GSDSTORE['WORKPATH']['INTERFACE']
                interFolderPath = f"{interfacePath}\\{subFolderPath}\\InterfaceData.yaml" if subFolderPath else f"{interfacePath}\\InterfaceData.yaml"
                if os.path.exists(interFolderPath):
                    if not self.__commonInterface_Object(interFolderPath): common.restore()

            '''元素节点'''
            if 'ELEMENT' in GSDSTORE['WORKPATH']:
                elementPath = GSDSTORE['WORKPATH']['ELEMENT']
                elementFolderPath = f"{elementPath}\\{subFolderPath}\\elementData.yaml" if subFolderPath else f"{elementPath}\\elementData.yaml"
                if not os.path.exists(elementFolderPath): elementFolderPath = None
            
            '''process'''
            try: 
                testCase_Object = excel_Read.TestCase()  # 实例化“用例步骤读取、操作”类对象
                testCase_Object.testCase(loopIndex, caseInformation, elementFolderPath)
            except Exception as error:
                print(traceback.format_exc()); break
        
        '''清除'''
        self.__initializeClear("suite_teardown", trusteeship.case_clear());common.restore()
    
    def __initializeClear(self, function, logEvent):
        suite = getattr(collector, function, None)
        if callable(suite):
            logEvent
            try:suite()
            except Exception as error: 
                CASESDATA[function] += 1; trusteeship.exceptinfo("blockage:");common.restore()