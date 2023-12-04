'''# Third_party'''
from openpyxl import load_workbook
from openpyxl.styles import *
'''# custom'''
from common import CASESDATA
from cases.read.run_Sheet_Row import RunSheetRow
from utils.public_Script import isLogicMethodClass
from utils.log.trusteeship import trusteeship


class TestCase:

    def __repr__(self) -> str:
        return "用例步骤文件操作控制类"

    def __init__(self) -> None:
        self.steps_Object = RunSheetRow()
        self.logicFunction_Object = isLogicMethodClass()
    
    def testCase(self, loopIndex, caseInformation_Json, caseNodeFile_Paht) -> None:
        """
         参数：
            @param loopIndex : 当前正在执行的用例文件索引标识
            @param caseInformation_Json : 用例步骤文件[Excel]文件信息
            @param aseNodeFile_Paht : 用例节点文件[CaseStep.yaml]
        """
        WB_Object = load_workbook(caseInformation_Json['caseFilePath'], read_only=True)
        elementExpression = self.logicFunction_Object.raw_YamlTxt(caseNodeFile_Paht) if caseNodeFile_Paht else None
        try:
            for sheetName in CASESDATA['toSheetNames'][loopIndex]:
                WS_Object = WB_Object[sheetName]
                WS_rowsMax = WS_Object.max_row
                caseImplement = ''  # 用例执行结果
                CASESDATA['implement'] += 1  # 执行用例数量+1
                caseName = WS_Object.cell(row=2, column=1).value; trusteeship.case_sheet(caseName); trusteeship.info(f'\n用例名：{caseName}------')  # 用例日志记录生成

                executeResult = self.steps_Object(WS_Object, WS_rowsMax, elementExpression)  # 单条用例步骤操作 [优先级 异常 > Fail > Pass]
                if executeResult['abnormalJudgment'] == True:
                    caseImplement = 'ABNORMAL'; CASESDATA['abnormalNum'] += 1
                    trusteeship.debug("Use case execution result record: ---------- ABNORMAL")
                elif False in executeResult['acceptResult']:
                    caseImplement = 'FAIL'; CASESDATA['failNum'] += 1
                    trusteeship.debug("Use case execution result record: ---------- Fail")
                else:
                    caseImplement = 'PASS'; CASESDATA['passNum'] += 1
                    trusteeship.debug("Use case execution result record: ---------- Pass")
                trusteeship.case_result(caseImplement)  # 用例执行结果记录
        finally:
            WB_Object.close()  # close