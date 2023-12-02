#!/usr/bin/env python
# coding: utf-8

from pathlib import Path
import pandas as pd
import openpyxl
from date_management import Date_Management


class Excel_Management:
    def __init__(self):
        pass

    def get_dataframe(Workbook: str, Sheet: str):
        dataframe = pd.DataFrame()
        try:
            dataframe = pd.read_excel(Path(Workbook), sheet_name=Sheet)
        except FileNotFoundError:
            error_msg = (
                "Error : File was not found at the "
                + Workbook[0: Workbook.rindex("\\") + 1]
                + " location."
            )
            print(f"{error_msg}")
        except ValueError:
            error_msg = f"Error : {Sheet} not found in the Workbook."
            print(f"{error_msg}")
        except Exception:
            error_msg = "Error : Unknown Error."
            print(f"{error_msg}")
        finally:
            if dataframe.empty:
                print("Null Dataframe loaded because of exception.")
            else:
                dataframe = Date_Management.timestamp_to_date(dataframe)
                print(f'''{Sheet} from {Workbook} loaded to dataframe
                      successfully.''')
            return dataframe

    def append_dataframe(Workbook: str, Sheet: str, dataframe):
        try:
            with pd.ExcelWriter(
                Path(Workbook), mode="a", engine="openpyxl",
                if_sheet_exists="overlay"
            ) as writer:
                dataframe.to_excel(
                    writer,
                    sheet_name=Sheet,
                    header=None,
                    startrow=writer.sheets[Sheet].max_row,
                    index=False,
                )
            print(f"Contents of dataframe written to {Sheet} in {Workbook}.")
        except KeyError:
            error_msg = "Error: Unknown Error."
            print(f"{error_msg}")
        except Exception:
            error_msg = "Error : Unknown Error."
            print(f"{error_msg}")

    def delete_excel_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        sheet_names_list = wb.sheetnames
        sheetno = sheet_names_list.index(Sheet)
        del wb[sheet_names_list[sheetno]]
        wb.save(Workbook)
        print(f"Sheet from {Workbook} deleted successfully.")

    def create_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.create_sheet(title=Sheet)
        wb.save(Workbook)
        print(f"{Sheet} created successfully in {Workbook}.")

    def overwrite_sheet(Workbook: str, Sheet: str, dataframe):
        header = dataframe.columns
        Excel_Management.delete_excel_sheet(Workbook=Workbook, Sheet=Sheet)
        Excel_Management.create_sheet(Workbook=Workbook, Sheet=Sheet)
        with pd.ExcelWriter(
            Path(Workbook), mode="a", engine="openpyxl",
            if_sheet_exists="overlay"
        ) as writer:
            dataframe.to_excel(
                writer, sheet_name=Sheet, header=header, startrow=0,
                index=False
            )
        print(f"Contents overwritten to {Sheet} in {Workbook}.")

    def reposition_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.move_sheet(Sheet, -(len(wb.sheetnames) - 1))
        wb.save(Workbook)
        wb.close()
        print("Sheet repositioned.")

    def excel_from_dataframe(Workbook: str, Sheet: str, Password: str, df):
        Excel_Management.remove_password_sheet(
            filepath=Workbook, sheetname=Sheet)
        with pd.ExcelWriter(
            Path(Workbook), mode="a", engine="openpyxl",
            if_sheet_exists="overlay"
        ) as writer:
            df.to_excel(
                writer,
                sheet_name=Sheet,
                header=None,
                startrow=writer.sheets[Sheet].max_row,
                index=False,
            )

        Excel_Management.add_sheet_password(
            filepath=Workbook, sheetname=Sheet, password=Password
        )
        print(f"Contents of dataframe written to {Workbook} in {Sheet}.")

    def remove_password_sheet(Workbook: str, Sheet: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.active = wb[Sheet]
        ws = wb.active
        ws.protection
        ws.protection.sheet = False
        wb.save(Workbook)
        wb.close()
        print("Password removed for editing.")

    def add_sheet_password(Workbook: str, Sheet: str, Password: str):
        wb = openpyxl.load_workbook(Workbook)
        wb.active = wb[Sheet]
        ws = wb.active
        ws.protection
        ws.protection.sheet = True
        ws.protection.password = Password
        wb.save(Workbook)
        wb.close()
        print("Password set after editing.")
