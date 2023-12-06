import re
import pathlib

from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles.alignment import horizontal_alignments, vertical_aligments


def create_new_excel():
    # 实例化
    wb = Workbook()
    return wb


def open_excel(filename):
    return load_workbook(filename=filename)


def save_excel(wb: Workbook, save_path=None):
    if save_path is None:
        save_path = pathlib.Path.home().joinpath("Downloads/output.xlsx")
    wb.save(save_path)


def create_new_sheet(wb: Workbook, sheet_name, index=None):
    return wb.create_sheet(
        title=sheet_name,
        # None代表在末尾加，否则就是index位置添加
        index=index
    )


def remove_sheet(wb: Workbook, sheet_name=None, sheet_index=None):
    if sheet_name is not None:
        wb.remove_sheet(worksheet=wb.get_sheet_by_name(sheet_name))
    elif sheet_index is not None:
        wb.remove_sheet(wb.worksheets[sheet_index])
    else:
        raise ValueError("sheet_name和sheet_index至少要有一个不为None")


def rename_sheet(ws: worksheet.worksheet.Worksheet, sheet_name):
    ws.title = sheet_name


def set_sheet_color(ws: worksheet.worksheet.Worksheet, aRGB ="00000000"):
    ws.sheet_properties.tabColor = colors.RgbColor(aRGB)


def write_data(ws: worksheet.worksheet.Worksheet, data):
    for row in data:
        ws.append(row)


def create_font_style(font_name="微软雅黑", font_size=12, font_color="00000000", bold=False, italic=False, vert_align=None, underline="none", strike=False):
    """ cell.font = return
        vertAlign: 垂直对齐方式, 'superscript', 'subscript', 'baseline', 前两个是上下标, 最后一个看不出来, 和正常的不一样
        underline: 下划线, 'single', 'double', 'singleAccounting', 'doubleAccounting', single是单下划线, double是双的, 带Accounting的比不带的要高一些, 估计是会计用的吧
        strike: 删除线
        font_color: aRGB, 十六进制
    """
    return Font(
        name=font_name,
        size=font_size,
        bold=bold,
        italic=italic,
        vertAlign=vert_align,
        underline=underline,
        strike=strike,
        color=font_color
    )


def create_fill_style(fill_type="none", start_color="FFFFFFFF", end_color="FF000000"):
    """ 单元格填充颜色, cell.fill = return
    Args:
        fill_type (_type_, optional): 各种花纹, 感觉solid就够用: ["none", "solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal", "darkTrellis", "darkUp", "darkVertical", "gray0625", "gray125", "lightDown", "lightGray", "lightGrid", "lightHorizontal", "lightTrellis", "lightUp", "lightVertical", "mediumGray"]
        start_color (str, optional): 默认是start_color, 但是有花纹的话end_color就用上了
        end_color (str, optional): 
    """
    __fill_types__ = ["none", "solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal", "darkTrellis", "darkUp", "darkVertical", "gray0625", "gray125", "lightDown", "lightGray", "lightGrid", "lightHorizontal", "lightTrellis", "lightUp", "lightVertical", "mediumGray"]
    if fill_type is not None:
        assert fill_type in __fill_types__, f"style应该是{', '.join(__fill_types__)}中的一种"

    return PatternFill(
        fill_type=fill_type,
        start_color=start_color,
        end_color=end_color
    )


def create_border_style(
        left_style=None, right_style=None, top_style=None, bottom_style=None, color="00000000"
        # 这些参数没看到效果先忽略
        # diagonal_style, diagonal_direction,
        # outline_style, vertical_style, horizontal_style
    ):
    """ cell.border = return
    style:  各种框框类型, 常用的是dashed/thin/thick, 'dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin'
    Args:
        left_style (_type_): _description_
        right_style (_type_): _description_
        top_style (_type_): _description_
        bottom_style (_type_): _description_
        diagonal_style (_type_): _description_
        diagonal_direction (_type_): _description_
        outline_style (_type_): _description_
        vertical_style (_type_): _description_
        horizontal_style (_type_): _description_

    Returns:
        _type_: _description_
    """

    __border_types__ = ('dashDot','dashDotDot', 'dashed','dotted', 'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot', 'mediumDashed', 'slantDashDot', 'thick', 'thin')
    for cur_style in [left_style, right_style, top_style, bottom_style]:
        if cur_style is not None:
            assert cur_style in __border_types__, f"style应该是{', '.join(__border_types__)}中的一种"
    
    return Border(
        left=Side(border_style=left_style, color=color),
        right=Side(border_style=right_style, color=color),
        top=Side(border_style=top_style, color=color),
        bottom=Side(border_style=bottom_style, color=color)
        # diagonal=Side(border_style=None, color='FF000000'),
        # diagonal_direction=0, 
        # outline=Side(border_style=None, color='FF000000'),
        # vertical=Side(border_style=None, color='FF000000'),
        # horizontal=Side(border_style=None, color='FF000000')
    )


def creaete_alignment_style(horizontal="general", vertical="center", text_rotation=0, wrap_text=True, shrink_to_fit=False, ident=0):
    """ cell.alignment = retrn
    
        horizontal_alignments = ("general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed", )
        vertical_aligments = ("top", "center", "bottom", "justify", "distributed")
    Args:
        horizontal (str, optional): 水平
        vertical (str, optional): 垂直
        text_rotation (int, optional): 旋转角度
        wrap_text (bool, optional): _description_. 换行
        shrink_to_fit (bool, optional): 似乎没用，是否缩小适应单元格
        ident (int, optional): 似乎没用

    """
    if horizontal is not None:
        assert horizontal in horizontal_alignments, f"horizontal应该是{', '.join(horizontal_alignments)}中的一种"
    if vertical is not None:
        assert vertical in vertical_aligments, f"vertical应该是{', '.join(vertical_aligments)}中的一种"
    
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        text_rotation=text_rotation,
        wrap_text=wrap_text,
        shrink_to_fit=shrink_to_fit,
        indent=ident
    )


def create_protection_style(locked=False, hidden=False):
    """
        锁定: 是不让改
        隐藏: 格式设置成";;;"之后，隐藏状态下公式编辑栏也看不到这个

        但是感觉隐藏没有卵用, 写个公式就能找出来了
    """
    return Protection(
        locked=locked,
        hidden=hidden
    )


def create_number_format():
    """
        Excel的数据格式, 比如百分比就是0.00%, 两位小数就是0.00等等
    """
    return "General"


def get_autojust_col_width(ws, n_cols):
    for col_idx in range(n_cols):
        col_dimension = ws.column_dimensions[chr(65 + col_idx)]
        col_dimension.width += 2


def get_autojust_col_width_bak(ws, start_row_idx, end_row_idx, col_idx=1):
    """本来用这个计算列的最大长度来控制列宽，但后来发现有自适应的方法

        Args:
            cur_str (_type_): _description_

        Returns:
            _type_: _description_
    """
    def get_length(cur_str):

        cur_str = str(cur_str)

        # 提取双字节字符的正则模板
        p = re.compile('[^\x00-\xff]+')  
        # 双字节字符数
        dualByteNum = len(''.join(p.findall(cur_str)))
        # 单字节字符数
        singleByteNum = len(cur_str) - dualByteNum
        
        return dualByteNum * 2 + singleByteNum

    col_max_length = 0

    for cur_row_idx in range(start_row_idx, end_row_idx + 1, 1):
        cur_cell = ws.cell(cur_row_idx, col_idx)
        cur_length = get_length(cur_cell.value)
        col_max_length = max(col_max_length, cur_length)


    return col_max_length + 2


    
