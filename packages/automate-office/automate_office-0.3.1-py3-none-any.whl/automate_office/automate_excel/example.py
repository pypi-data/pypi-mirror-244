import pathlib
import datetime

import numpy as np
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from automate_office.automate_excel.excel_funcs import *


def get_border_style(ws, cur_row_idx, cur_col_idx, border_color="00FF6238"):
    left_style, right_style, top_style, bottom_style = ["dashed"] * 4
    if cur_row_idx == ws.min_row:
        top_style = "thin"
    if cur_row_idx == ws.max_row:
        bottom_style = "thin"
    if cur_col_idx == ws.min_column:
        left_style = "thin"
    if cur_col_idx == ws.max_column:
        right_style = "thin"
    return create_border_style(right_style=right_style, left_style=left_style, bottom_style=bottom_style, top_style=top_style, color=border_color)


def get_number_format(ws, cur_row_idx, cur_col_idx):
    if cur_row_idx == ws.min_row:
        return "General"
    
    if cur_col_idx == 2:
        return "0.00%"
    elif cur_col_idx == 3:
        return "#,##0.00;[红色]-#,##0.00"
    elif cur_col_idx == 4:
        # 中文在自适应列宽的时候似乎总会短一截
        return 'yyyy"年"m"月"d"日"'
    else:
        return "General"
    

if __name__ == "__main__":  
    file_path = pathlib.Path.home().joinpath("Downloads/format.xlsx")
    sheet_name = "Format"

    df = pd.DataFrame(
        {
            "col1": list("abcdefg"),
            "col2": np.random.random(size=(7, )),
            "col3": np.random.randint(low=-1e5, high=1e5, size=(7, )),
            "col4": pd.date_range(start=datetime.datetime(2023, 12, 1).date(), end=datetime.datetime(2023, 12, 7).date())
        }
    )
    n_rows, n_cols = df.shape[0] + 1, df.shape[1]
    
    # 格式化
    wb = create_new_excel()
    ws = wb.active
    rename_sheet(ws, sheet_name=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # # 注意遍历的方法
    # # pandas样式
    # for cell in ws['A'] + ws[1]:
    #     cell.style = 'Pandas'
    # wb.save("pandas_openpyxl.xlsx")

    alignment = creaete_alignment_style(horizontal="center", vertical="center")
    title_font = create_font_style(font_name="微软雅黑", font_size=16, font_color="00FFFFFF", bold=True)
    content_font = create_font_style(font_name="微软雅黑", font_size=12, font_color="00000000")
    title_fill = create_fill_style(fill_type="solid", start_color="00FF6238")

    # 遍历单元格设置格式
    for row_cells in ws.iter_rows():
        for cur_cell in row_cells:
            if cur_cell.row == 1:
                cur_cell.font = title_font
                cur_cell.fill = title_fill
            else:
                cur_cell.font = content_font

            cur_cell.alignment = alignment
            cur_cell.border = get_border_style(ws=ws, cur_col_idx=cur_cell.column, cur_row_idx=cur_cell.row)
            cur_cell.number_format = get_number_format(ws, cur_col_idx=cur_cell.column, cur_row_idx=cur_cell.row)

    # 自适应列宽
    get_autojust_col_width(ws, n_cols)

    wb.save(file_path)
